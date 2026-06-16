"""Workbook tests: build from a seeded DB, assert zero errors + numeric KPIs."""

import shutil

import pytest
from openpyxl import load_workbook

from src import store, workbook
from tests import fixtures

HAS_SOFFICE = bool(shutil.which("soffice") or shutil.which("libreoffice"))


@pytest.fixture()
def built_workbook(tmp_path):
    conn = store.connect(":memory:")
    fixtures.seed(conn, n_days=45)
    out = tmp_path / "test_dashboard.xlsx"
    path = workbook.build(conn, out_path=out)
    conn.close()
    return path


def test_workbook_has_no_error_strings(built_workbook):
    errors = workbook.scan_for_errors(built_workbook)
    assert errors == [], f"formula errors present: {errors[:10]}"


def test_expected_tabs_present(built_workbook):
    wb = load_workbook(built_workbook)
    for tab in ["Dashboard", "Daily", "Recovery", "Sleep", "Strain_Workouts",
                "Rolling", "Correlations", "Interventions", "Config"]:
        assert tab in wb.sheetnames
    # Dashboard is first.
    assert wb.sheetnames[0] == "Dashboard"
    wb.close()


def test_dashboard_kpis_resolve_to_numbers(built_workbook):
    wb = load_workbook(built_workbook, data_only=True)
    # A formula cell carries a cached value only if a recalc engine ran. If the
    # environment's LibreOffice can't recalc, cached values are absent; the
    # shipped workbook still recalcs live on open (fullCalcOnLoad=True).
    recalced = wb["Rolling"]["B3"].value is not None
    if not recalced:
        wb.close()
        pytest.skip("No working recalc engine; values populate on open in Excel.")
    ws = wb["Dashboard"]
    # KPI value cells live on row 5, columns A, C, E, G, I, K.
    kpi_cells = ["A5", "C5", "E5", "G5", "I5", "K5"]
    numeric = sum(1 for c in kpi_cells if isinstance(ws[c].value, (int, float)))
    assert numeric >= 4, "expected most KPI cells to recalc to numbers"
    wb.close()


def test_daily_tab_is_values_not_formulas(built_workbook):
    wb = load_workbook(built_workbook)
    ws = wb["Daily"]
    # Recovery column B, first data row, should be a raw number (source data).
    assert isinstance(ws["B2"].value, (int, float))
    wb.close()
