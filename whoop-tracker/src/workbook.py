"""Build the dashboard-grade whoop_dashboard.xlsx.

Design principles enforced here:
  * Data tabs (Daily, Recovery, Sleep, Strain_Workouts, Interventions) hold
    values. EVERY analytical number (Rolling stats, z-scores, correlations,
    KPI baselines, threshold colouring) is an Excel FORMULA referencing those
    tabs, so the workbook recalculates live when raw tabs are refreshed.
  * All divisions/lookups are wrapped in IFERROR so the workbook opens with
    ZERO formula errors (#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?, ...).
  * Native Excel sparklines are injected as x14 extension XML after build,
    since openpyxl has no sparkline writer.

After writing, ``build`` runs a LibreOffice headless recalc pass to populate
cached values, then scans every cell for error strings and raises if any are
found.
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config
from . import analytics, store, transform

log = logging.getLogger("whoop.workbook")

# ---- styling -------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color="1F2A44")
SUB_FONT = Font(name=FONT_NAME, italic=True, size=9, color="666666")
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10, color="1F2A44")
INPUT_FONT = Font(name=FONT_NAME, color="0000CC")  # blue = input convention
BASE_FONT = Font(name=FONT_NAME, size=10)
KPI_VALUE_FONT = Font(name=FONT_NAME, bold=True, size=22, color="1F2A44")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

RED = "F8696B"
YELLOW = "FFEB84"
GREEN = "63BE7B"

ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#####")


# --------------------------------------------------------------------------
# Small helpers
# --------------------------------------------------------------------------
def _col(idx: int) -> str:
    return get_column_letter(idx)


def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autowidth(ws, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = None
        for cell in col_cells:
            letter = cell.column_letter
            v = cell.value
            if v is not None:
                length = max(length, len(str(v)))
        if letter:
            ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


# --------------------------------------------------------------------------
# Data tabs
# --------------------------------------------------------------------------
def _write_daily(wb: Workbook, daily: list[dict], items: list[str]):
    ws = wb.create_sheet("Daily")
    headers = [h for _, h in transform.DAILY_COLUMNS] + items
    ws.append(headers)
    keys = [k for k, _ in transform.DAILY_COLUMNS]

    for row in daily:
        values = []
        for k in keys:
            values.append(row.get(k))
        for it in items:
            present = any(iv.get("item") == it for iv in row.get("interventions", []))
            values.append(1 if present else 0)
        ws.append(values)

    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    last_row = len(daily) + 1

    # Conditional colour scales on the analytical columns.
    if last_row >= 2:
        scales = {
            "B": (RED, YELLOW, GREEN),   # recovery (low bad)
            "C": (RED, YELLOW, GREEN),   # HRV (low bad)
            "D": (GREEN, YELLOW, RED),   # RHR (low good)
            "E": (GREEN, YELLOW, RED),   # strain (informational; high = more load)
            "H": (RED, YELLOW, GREEN),   # hours slept (low bad)
            "F": (RED, YELLOW, GREEN),   # sleep perf (low bad)
        }
        for letter, (lo, mid, hi) in scales.items():
            rng = f"{letter}2:{letter}{last_row}"
            ws.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min", start_color=lo,
                    mid_type="percentile", mid_value=50, mid_color=mid,
                    end_type="max", end_color=hi,
                ),
            )
    _autowidth(ws)
    return ws, last_row


def _write_recovery(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Recovery")
    headers = ["Cycle ID", "Sleep ID", "State", "Recovery %", "RHR", "HRV (ms)",
               "SpO2 %", "Skin °C", "Skin °F (calc)", "Calibrating"]
    ws.append(headers)
    for r in sorted(rows, key=lambda x: x.get("cycle_id") or 0, reverse=True):
        ws.append([r.get("cycle_id"), r.get("sleep_id"), r.get("score_state"),
                   r.get("recovery_score"), r.get("resting_hr"), r.get("hrv_rmssd_ms"),
                   r.get("spo2_pct"), r.get("skin_temp_c"), None, r.get("user_calibrating")])
    for i in range(2, len(rows) + 2):
        ws.cell(row=i, column=9).value = f'=IFERROR(H{i}*9/5+32,"")'
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    _autowidth(ws)


def _write_sleep(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Sleep")
    headers = ["Sleep ID", "Start", "End", "Nap", "State", "Perf %", "Consistency %",
               "Efficiency %", "In Bed (ms)", "Awake (ms)", "Light (ms)", "SWS (ms)",
               "REM (ms)", "Resp Rate", "Need (ms)", "In Bed (h, calc)",
               "Asleep (h, calc)", "Need (h, calc)"]
    ws.append(headers)
    for r in sorted(rows, key=lambda x: x.get("start") or "", reverse=True):
        ws.append([r.get("id"), r.get("start"), r.get("end"), r.get("nap"),
                   r.get("score_state"), r.get("sleep_perf_pct"),
                   r.get("sleep_consistency_pct"), r.get("sleep_efficiency_pct"),
                   r.get("total_in_bed_ms"), r.get("total_awake_ms"),
                   r.get("total_light_ms"), r.get("total_sws_ms"), r.get("total_rem_ms"),
                   r.get("respiratory_rate"), r.get("sleep_need_ms"), None, None, None])
    for i in range(2, len(rows) + 2):
        ws.cell(row=i, column=16).value = f'=IFERROR(I{i}/3600000,"")'
        ws.cell(row=i, column=17).value = f'=IFERROR((K{i}+L{i}+M{i})/3600000,"")'
        ws.cell(row=i, column=18).value = f'=IFERROR(O{i}/3600000,"")'
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    _autowidth(ws)


def _write_workouts(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Strain_Workouts")
    headers = ["Workout ID", "Start", "End", "Sport", "State", "Strain", "Avg HR",
               "Max HR", "Kilojoule", "Calories (calc)", "Distance (m)", "Alt Gain (m)"]
    ws.append(headers)
    for r in sorted(rows, key=lambda x: x.get("start") or "", reverse=True):
        sport = r.get("sport_name") or r.get("sport_id")
        ws.append([r.get("id"), r.get("start"), r.get("end"), sport, r.get("score_state"),
                   r.get("strain"), r.get("avg_hr"), r.get("max_hr"), r.get("kilojoule"),
                   None, r.get("distance_m"), r.get("altitude_gain_m")])
    for i in range(2, len(rows) + 2):
        ws.cell(row=i, column=10).value = f'=IFERROR(I{i}*{config.KJ_TO_CAL},"")'
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    _autowidth(ws)


def _write_interventions(wb: Workbook, daily: list[dict]):
    ws = wb.create_sheet("Interventions")
    headers = ["date", "category", "item", "dose", "notes"]
    ws.append(headers)
    seen = []
    for row in daily:
        for iv in row.get("interventions", []):
            seen.append([row["date"], iv.get("category"), iv.get("item"),
                         iv.get("dose"), iv.get("notes")])
    # Fall back to the CSV directly so the tab mirrors the editable log even if
    # some logged dates have no matching cycle yet.
    if not seen:
        for d, ivs in sorted(transform.load_interventions().items()):
            for iv in ivs:
                seen.append([d, iv.get("category"), iv.get("item"),
                             iv.get("dose"), iv.get("notes")])
    for r in seen:
        ws.append(r)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    # Data-validation dropdown on `category`.
    cats = "supplement,nutrition,training,sleep_protocol,substance,environment"
    dv = DataValidation(type="list", formula1=f'"{cats}"', allow_blank=True)
    dv.add(f"B2:B{max(len(seen) + 50, 200)}")
    ws.add_data_validation(dv)
    _autowidth(ws)


# --------------------------------------------------------------------------
# Formula tabs
# --------------------------------------------------------------------------
# Daily column letters for the six tracked metrics.
_METRIC_COLS = {
    "recovery": "B",
    "hrv": "C",
    "rhr": "D",
    "strain": "E",
    "hours": "H",
    "sleep_perf": "F",
}
_METRIC_LABELS = {
    "recovery": "Recovery %",
    "hrv": "HRV (ms)",
    "rhr": "RHR (bpm)",
    "strain": "Day Strain",
    "hours": "Hours Slept",
    "sleep_perf": "Sleep Perf %",
}
_METRIC_ORDER = ["recovery", "hrv", "rhr", "strain", "hours", "sleep_perf"]


def _rolling_cols(metric_index: int) -> dict:
    """Return Rolling-tab column letters for a metric's 5 stat columns."""
    base = 2 + metric_index * 5  # col A is Date
    return {
        "mean7": _col(base),
        "mean30": _col(base + 1),
        "sd30": _col(base + 2),
        "baseline": _col(base + 3),
        "z": _col(base + 4),
    }


def _write_rolling(wb: Workbook, last_row: int):
    ws = wb.create_sheet("Rolling")
    # Two header rows: metric group + stat name.
    ws.cell(row=1, column=1, value="").font = LABEL_FONT
    ws.cell(row=2, column=1, value="Date")
    for mi, key in enumerate(_METRIC_ORDER):
        cols = _rolling_cols(mi)
        ws.cell(row=1, column=2 + mi * 5, value=_METRIC_LABELS[key])
        for name, letter in (("7d Mean", cols["mean7"]), ("30d Mean", cols["mean30"]),
                             ("30d SD", cols["sd30"]), ("Baseline", cols["baseline"]),
                             ("Z-score", cols["z"])):
            ws[f"{letter}2"] = name

    ncols = 1 + len(_METRIC_ORDER) * 5
    _style_header_row(ws, 1, ncols)
    _style_header_row(ws, 2, ncols)
    ws.freeze_panes = "B3"

    # One row per Daily day (data starts at Daily row 2 -> Rolling row 3).
    for r in range(2, last_row + 1):
        out_row = r + 1  # offset by the extra header row
        ws.cell(row=out_row, column=1, value=f"=Daily!A{r}")
        end7 = min(r + 6, last_row)
        end30 = min(r + 29, last_row)
        for mi, key in enumerate(_METRIC_ORDER):
            dcol = _METRIC_COLS[key]
            rc = _rolling_cols(mi)
            rng7 = f"Daily!{dcol}{r}:{dcol}{end7}"
            rng30 = f"Daily!{dcol}{r}:{dcol}{end30}"
            cal30 = f"Daily!$R{r}:$R{end30}"
            ws[f"{rc['mean7']}{out_row}"] = f'=IFERROR(AVERAGE({rng7}),"")'
            ws[f"{rc['mean30']}{out_row}"] = f'=IFERROR(AVERAGE({rng30}),"")'
            ws[f"{rc['sd30']}{out_row}"] = f'=IFERROR(STDEV({rng30}),"")'
            # Baseline excludes calibrating days (flag column R == 0).
            ws[f"{rc['baseline']}{out_row}"] = (
                f'=IFERROR(AVERAGEIFS({rng30},{cal30},0),"")'
            )
            bcell = f"{rc['baseline']}{out_row}"
            scell = f"{rc['sd30']}{out_row}"
            vcell = f"Daily!{dcol}{r}"
            ws[f"{rc['z']}{out_row}"] = (
                f'=IFERROR(({vcell}-{bcell})/{scell},"")'
            )
    _autowidth(ws, max_width=14)
    return ws


# Predictors for the correlation matrix: (label, Daily column letter).
def _correlation_predictors(items: list[str]) -> list[tuple[str, str]]:
    base = [
        ("Prior-day Strain", "E"),
        ("Sleep Hours", "H"),
        ("Sleep Perf %", "F"),
        ("Sleep Consistency %", "G"),
        ("Bedtime (h)", "O"),
        ("Respiratory Rate", "K"),
    ]
    # Intervention flag columns start after the 18 fixed Daily columns.
    for i, it in enumerate(items):
        base.append((f"flag: {it}", _col(19 + i)))
    return base


def _write_correlations(wb: Workbook, last_row: int, items: list[str]):
    ws = wb.create_sheet("Correlations")
    preds = _correlation_predictors(items)

    ws["A1"] = "Drivers of Next-Morning Recovery & HRV (Pearson r, prior day → next day)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Predictor"; ws["B2"] = "r vs Recovery"; ws["C2"] = "n"
    ws["D2"] = "r vs HRV"; ws["E2"] = "n"; ws["F2"] = "|r| Recovery"
    _style_header_row(ws, 2, 6)

    # Offset ranges: predictor on day d (rows 3..last) paired with outcome on
    # day d+1 (rows 2..last-1), because Daily is newest-first.
    have_pairs = last_row >= 3
    first_pred_row = 3
    for i, (label, dcol) in enumerate(preds):
        r = 3 + i
        ws[f"A{r}"] = label
        if have_pairs:
            pred_rng = f"Daily!{dcol}{3}:{dcol}{last_row}"
            rec_rng = f"Daily!B2:B{last_row - 1}"
            hrv_rng = f"Daily!C2:C{last_row - 1}"
            ws[f"B{r}"] = f'=IFERROR(CORREL({pred_rng},{rec_rng}),"")'
            ws[f"C{r}"] = f'=COUNT({pred_rng})'
            ws[f"D{r}"] = f'=IFERROR(CORREL({pred_rng},{hrv_rng}),"")'
            ws[f"E{r}"] = f'=COUNT({pred_rng})'
            ws[f"F{r}"] = f'=IFERROR(ABS(B{r}),"")'
        else:
            for cletter in ("B", "C", "D", "E", "F"):
                ws[f"{cletter}{r}"] = ""
    last_pred_row = 3 + len(preds) - 1

    # ---- Strain-quartile -> next-day recovery -------------------------------
    qstart = last_pred_row + 3
    ws[f"A{qstart}"] = "Strain Quartile → Mean Next-Day Recovery"
    ws[f"A{qstart}"].font = TITLE_FONT
    hdr = qstart + 1
    ws[f"A{hdr}"] = "Bucket"; ws[f"B{hdr}"] = "Strain ≥"; ws[f"C{hdr}"] = "Strain <"
    ws[f"D{hdr}"] = "Mean Next-Day Recovery"; ws[f"E{hdr}"] = "n"
    _style_header_row(ws, hdr, 5)

    # Helper pair block (far right): strain(d) & next-day recovery(d+1).
    hcol_s, hcol_r = "O", "P"
    ws[f"{hcol_s}1"] = "Strain(d)"
    ws[f"{hcol_r}1"] = "NextRec(d+1)"
    if have_pairs:
        for r in range(first_pred_row, last_row + 1):
            ws[f"{hcol_s}{r}"] = f"=Daily!E{r}"
            ws[f"{hcol_r}{r}"] = f"=Daily!B{r - 1}"
    srng = f"{hcol_s}{first_pred_row}:{hcol_s}{last_row}"
    rrng = f"{hcol_r}{first_pred_row}:{hcol_r}{last_row}"

    labels = ["Q1 (low)", "Q2", "Q3", "Q4 (high)"]
    for b in range(4):
        r = hdr + 1 + b
        ws[f"A{r}"] = labels[b]
        if have_pairs:
            lo_q = b      # QUARTILE index for lower edge
            hi_q = b + 1
            ws[f"B{r}"] = f'=IFERROR(QUARTILE({srng},{lo_q}),"")'
            ws[f"C{r}"] = f'=IFERROR(QUARTILE({srng},{hi_q}),"")'
            op_hi = "<=" if b == 3 else "<"
            ws[f"D{r}"] = (
                f'=IFERROR(AVERAGEIFS({rrng},{srng},">="&B{r},{srng},"{op_hi}"&C{r}),"")'
            )
            ws[f"E{r}"] = (
                f'=COUNTIFS({srng},">="&B{r},{srng},"{op_hi}"&C{r})'
            )
        else:
            for cletter in ("B", "C", "D", "E"):
                ws[f"{cletter}{r}"] = ""

    ws.column_dimensions["A"].width = 26
    for c in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[c].width = 16
    # Stash range metadata for the Dashboard "what's driving" lookup.
    ws._pred_name_range = f"Correlations!$A${first_pred_row}:$A${last_pred_row}"
    ws._pred_r_range = f"Correlations!$B${first_pred_row}:$B${last_pred_row}"
    ws._pred_abs_range = f"Correlations!$F${first_pred_row}:$F${last_pred_row}"
    return ws


def _write_config(wb: Workbook):
    ws = wb.create_sheet("Config")
    ws["A1"] = "WHOOP Dashboard — Configuration"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Blue cells are inputs you can retune; formulas elsewhere reference them."
    ws["A2"].font = SUB_FONT

    rows = [
        ("Recovery red max (0–N = red)", config.RECOVERY_RED_MAX, "recovery_red"),
        ("Recovery yellow max (N = yellow; above = green)", config.RECOVERY_YELLOW_MAX, "recovery_yellow"),
        ("Z-score anomaly threshold (|z| ≥)", config.Z_ANOMALY, "z_anomaly"),
        ("Kilojoule → calories factor", config.KJ_TO_CAL, "kj_to_cal"),
    ]
    ws["A4"] = "Setting"; ws["B4"] = "Value"
    _style_header_row(ws, 4, 2)
    r = 5
    named = {}
    for label, value, name in rows:
        ws[f"A{r}"] = label
        cell = ws[f"B{r}"]
        cell.value = value
        cell.font = INPUT_FONT
        named[name] = f"Config!$B${r}"
        r += 1

    r += 1
    ws[f"A{r}"] = "Cycle → calendar-day rule"; ws[f"A{r}"].font = LABEL_FONT
    ws[f"A{r+1}"] = config.CYCLE_DAY_RULE
    ws[f"A{r+1}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=8)

    r += 5
    ws[f"A{r}"] = "Color thresholds (WHOOP style)"; ws[f"A{r}"].font = LABEL_FONT
    ws[f"A{r+1}"] = "Recovery 0–33 red · 34–66 yellow · 67–100 green"
    ws[f"A{r+2}"] = "HRV / Sleep / Recovery: higher = better. RHR: lower = better."

    r += 4
    ws[f"A{r}"] = "Units"; ws[f"A{r}"].font = LABEL_FONT
    ws[f"A{r+1}"] = "Durations → hours · kilojoule → kcal (×0.239006) · temp °C (+ derived °F)"

    r += 3
    ws[f"A{r}"] = "Refresh instructions"; ws[f"A{r}"].font = LABEL_FONT
    ws[f"A{r+1}"] = "Run `whoop all` (sync + build) to regenerate, or `whoop sync` then `whoop build`."
    ws[f"A{r+2}"] = "All analytical tabs recalc live from the Daily tab on open."

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 16
    return named


def _write_dashboard(wb: Workbook, daily: list[dict], last_row: int, cfg: dict, corr_ws):
    ws = wb.create_sheet("Dashboard", 0)  # first tab
    ws.sheet_view.showGridLines = False
    ws["A1"] = "WHOOP Daily Cockpit"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = '=IF(COUNT(Daily!A:A)=0,"No data yet","Latest day: "&TEXT(Daily!A2,"yyyy-mm-dd"))'
    ws["A2"].font = SUB_FONT

    # ---- KPI band -----------------------------------------------------------
    kpi_defs = [
        ("Recovery %", "recovery"),
        ("HRV (ms)", "hrv"),
        ("RHR (bpm)", "rhr"),
        ("Day Strain", "strain"),
        ("Hours Slept", "hours"),
        ("Sleep Perf %", "sleep_perf"),
    ]
    header_row = 4
    val_row = 5
    base_row = 6
    z_row = 7
    for i, (label, key) in enumerate(kpi_defs):
        col = 1 + i * 2  # 2-column spacing per KPI
        cl = _col(col)
        cl2 = _col(col + 1)
        ws.merge_cells(f"{cl}{header_row}:{cl2}{header_row}")
        h = ws[f"{cl}{header_row}"]
        h.value = label
        h.font = LABEL_FONT
        h.alignment = CENTER

        dcol = _METRIC_COLS[key]
        ws.merge_cells(f"{cl}{val_row}:{cl2}{val_row}")
        v = ws[f"{cl}{val_row}"]
        v.value = f'=IFERROR(Daily!{dcol}2,"—")'
        v.font = KPI_VALUE_FONT
        v.alignment = CENTER
        v.border = BORDER

        mi = _METRIC_ORDER.index(key)
        rc = _rolling_cols(mi)
        ws.merge_cells(f"{cl}{base_row}:{cl2}{base_row}")
        b = ws[f"{cl}{base_row}"]
        b.value = f'=IFERROR("base "&TEXT(Rolling!{rc["baseline"]}3,"0.0"),"base —")'
        b.font = SUB_FONT
        b.alignment = CENTER

        ws.merge_cells(f"{cl}{z_row}:{cl2}{z_row}")
        z = ws[f"{cl}{z_row}"]
        z.value = f'=IFERROR("z="&TEXT(Rolling!{rc["z"]}3,"0.00"),"z=—")'
        z.font = SUB_FONT
        z.alignment = CENTER

    # Recovery KPI traffic-light fill, driven by Config thresholds.
    rec_cell = f"A{val_row}"
    ws.conditional_formatting.add(
        rec_cell,
        FormulaRule(formula=[f"AND(ISNUMBER(Daily!B2),Daily!B2<={cfg['recovery_red']})"],
                    fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        rec_cell,
        FormulaRule(formula=[f"AND(ISNUMBER(Daily!B2),Daily!B2<={cfg['recovery_yellow']})"],
                    fill=PatternFill("solid", fgColor=YELLOW)),
    )
    ws.conditional_formatting.add(
        rec_cell,
        FormulaRule(formula=[f"AND(ISNUMBER(Daily!B2),Daily!B2>{cfg['recovery_yellow']})"],
                    fill=PatternFill("solid", fgColor=GREEN)),
    )

    # ---- Trend block (labels; sparklines injected post-build) ---------------
    trend_row = 10
    ws[f"A{trend_row}"] = "30-Day Trends"
    ws[f"A{trend_row}"].font = LABEL_FONT
    spark_specs = []
    spark_color = {"recovery": "FF63BE7B", "hrv": "FF1F77B4", "rhr": "FFD62728",
                   "strain": "FFFF7F0E", "hours": "FF9467BD"}
    spark_metrics = [("Recovery", "recovery"), ("HRV", "hrv"), ("RHR", "rhr"),
                     ("Day Strain", "strain"), ("Hours Slept", "hours")]
    for j, (label, key) in enumerate(spark_metrics):
        r = trend_row + 1 + j
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = BASE_FONT
        dcol = _METRIC_COLS[key]
        spark_end = min(last_row, 31)
        rng = f"Daily!{dcol}2:{dcol}{spark_end}"
        # sparkline drawn in column B..D area for this row
        ws.merge_cells(f"B{r}:E{r}")
        spark_specs.append((rng, f"B{r}", spark_color[key]))

    # ---- Flags today --------------------------------------------------------
    flag_row = trend_row
    ws[f"G{flag_row}"] = "Flags Today (|z| ≥ threshold)"
    ws[f"G{flag_row}"].font = LABEL_FONT
    for j, key in enumerate(_METRIC_ORDER):
        r = flag_row + 1 + j
        mi = _METRIC_ORDER.index(key)
        rc = _rolling_cols(mi)
        zc = f"Rolling!{rc['z']}3"
        ws.merge_cells(f"G{r}:I{r}")
        ws[f"G{r}"] = (
            f'=IFERROR(IF(ABS({zc})>={cfg["z_anomaly"]},'
            f'"⚠ {_METRIC_LABELS[key]}  z="&TEXT({zc},"0.0"),""),"")'
        )
        ws[f"G{r}"].font = BASE_FONT

    # ---- What's driving recovery (top 5 |correlation|) ----------------------
    drive_row = 18
    ws[f"A{drive_row}"] = "What's Driving Recovery (top correlations)"
    ws[f"A{drive_row}"].font = LABEL_FONT
    ws[f"A{drive_row+1}"] = "Driver"; ws[f"C{drive_row+1}"] = "r"
    ws[f"A{drive_row+1}"].font = HEADER_FONT
    ws[f"A{drive_row+1}"].fill = HEADER_FILL
    name_rng = corr_ws._pred_name_range
    r_rng = corr_ws._pred_r_range
    abs_rng = corr_ws._pred_abs_range
    for k in range(1, 6):
        r = drive_row + 1 + k
        kth = f'LARGE({abs_rng},{k})'
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"] = (
            f'=IFERROR(INDEX({name_rng},MATCH({kth},{abs_rng},0)),"—")'
        )
        ws[f"C{r}"] = (
            f'=IFERROR(INDEX({r_rng},MATCH({kth},{abs_rng},0)),"")'
        )
        ws[f"A{r}"].font = BASE_FONT
        ws[f"C{r}"].font = BASE_FONT

    # widths
    for c in range(1, 12):
        ws.column_dimensions[_col(c)].width = 13
    ws.column_dimensions["A"].width = 18
    return ws, spark_specs


def _write_raw(wb: Workbook, conn):
    """Hidden audit tabs holding raw JSON per resource."""
    specs = [("Raw_Cycles", "cycles", "id"), ("Raw_Recovery", "recovery", "cycle_id"),
             ("Raw_Sleep", "sleep", "id"), ("Raw_Workouts", "workouts", "id")]
    for title, table, key in specs:
        ws = wb.create_sheet(title)
        ws.append([key, "raw_json"])
        for row in store.fetch_all(conn, table):
            ws.append([row.get(key), row.get("raw_json")])
        _style_header_row(ws, 1, 2)
        ws.sheet_state = "hidden"
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 80


# --------------------------------------------------------------------------
# Sparkline injection + recalc + validation
# --------------------------------------------------------------------------
def _inject_sparklines(path: Path, sheet_title: str, specs: list[tuple]) -> None:
    """Insert native x14 sparkline groups into `sheet_title` of the xlsx."""
    if not specs:
        return
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
        rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        contents = {n: zf.read(n) for n in names}

    # Map sheet title -> r:id -> worksheet target.
    m = re.search(rf'<sheet[^>]*name="{re.escape(sheet_title)}"[^>]*r:id="([^"]+)"', wb_xml)
    if not m:
        log.warning("Sparkline inject: sheet %s not found.", sheet_title)
        return
    rid = m.group(1)
    # Build {Id: Target} for all relationships; attribute order is not fixed.
    id_to_target = {}
    for rel in re.findall(r"<Relationship\b[^>]*/?>", rels):
        rid_m = re.search(r'Id="([^"]+)"', rel)
        tgt_m = re.search(r'Target="([^"]+)"', rel)
        if rid_m and tgt_m:
            id_to_target[rid_m.group(1)] = tgt_m.group(1)
    target = id_to_target.get(rid)
    if not target:
        log.warning("Sparkline inject: no rels target for %s.", rid)
        return
    # Normalize to a package-root path (handles '/xl/...', 'xl/...', 'worksheets/...').
    if target.startswith("/"):
        target = target.lstrip("/")
    elif not target.startswith("xl/"):
        target = "xl/" + target

    xml = contents[target].decode("utf-8")

    groups = []
    for rng, cell, color in specs:
        groups.append(
            '<x14:sparklineGroup type="line" displayEmptyCellsAs="gap" '
            'lineWeight="1.5" markers="0">'
            f'<x14:colorSeries rgb="{color}"/>'
            f'<x14:colorAxis rgb="FF000000"/>'
            f'<x14:colorMarkers rgb="{color}"/>'
            '<x14:sparklines><x14:sparkline>'
            f'<xm:f>{rng}</xm:f><xm:sqref>{cell}</xm:sqref>'
            '</x14:sparkline></x14:sparklines></x14:sparklineGroup>'
        )
    ext = (
        '<extLst><ext '
        'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
        'uri="{05C60535-1F16-4fd2-B633-F4F36F0B64E0}">'
        '<x14:sparklineGroups '
        'xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
        + "".join(groups) +
        '</x14:sparklineGroups></ext></extLst>'
    )

    if "<extLst>" in xml:
        xml = xml.replace("<extLst>", "<extLst>" + ext.replace("<extLst>", "").replace("</extLst>", ""), 1)
    else:
        xml = xml.replace("</worksheet>", ext + "</worksheet>")
    contents[target] = xml.encode("utf-8")

    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in names:
            zf.writestr(n, contents[n])
    os.replace(tmp, path)
    log.info("Injected %d sparkline(s) into %s.", len(specs), sheet_title)


# LibreOffice profile that forces "always recalculate" on load so the
# converted file carries fresh cached values (0 = Always).
_RECALC_XCU = (
    '<?xml version="1.0" encoding="UTF-8"?>'
    '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
    'xmlns:xs="http://www.w3.org/2001/XMLSchema">'
    '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>'
    '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>'
    '</oor:items>'
)


def _recalc(path: Path) -> bool:
    """Recalculate cached values with LibreOffice headless. Returns success."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        log.warning("LibreOffice not found; skipping recalc (cached values absent).")
        return False
    with tempfile.TemporaryDirectory() as tmp:
        profile = Path(tmp) / "profile"
        user_dir = profile / "user"
        user_dir.mkdir(parents=True, exist_ok=True)
        (user_dir / "registrymodifications.xcu").write_text(_RECALC_XCU)
        try:
            subprocess.run(
                [soffice, "--headless", "--nologo", "--nofirststartwizard",
                 f"-env:UserInstallation=file://{profile}",
                 "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                 "--outdir", tmp, str(path)],
                check=True, capture_output=True, timeout=180,
                env={**os.environ, "HOME": tmp},
            )
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
            log.warning("LibreOffice recalc failed: %s", exc)
            return False
        produced = Path(tmp) / path.name
        if produced.exists():
            shutil.copy(produced, path)
            return True
    return False


def scan_for_errors(path: Path) -> list[str]:
    """Return a list of 'Sheet!Cell=value' strings for any error cells."""
    from openpyxl import load_workbook

    found = []
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(tok in v for tok in ERROR_TOKENS):
                    found.append(f"{ws.title}!{cell.coordinate}={v}")
    wb.close()
    return found


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------
def build(conn, out_path: Optional[Path] = None, recalc: bool = True) -> Path:
    out_path = Path(out_path or config.WORKBOOK_PATH)

    cycles = store.fetch_all(conn, "cycles")
    recovery = store.fetch_all(conn, "recovery")
    sleep = store.fetch_all(conn, "sleep")
    workouts = store.fetch_all(conn, "workouts")
    interventions = transform.load_interventions()
    daily = transform.build_daily(cycles, recovery, sleep, workouts, interventions)
    items = transform.all_intervention_items(daily)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    _, last_row = _write_daily(wb, daily, items)
    _write_recovery(wb, recovery)
    _write_sleep(wb, sleep)
    _write_workouts(wb, workouts)
    _write_interventions(wb, daily)
    _write_rolling(wb, last_row)
    corr_ws = _write_correlations(wb, last_row, items)
    cfg = _write_config(wb)
    _, spark_specs = _write_dashboard(wb, daily, last_row, cfg, corr_ws)
    _write_raw(wb, conn)

    # Order: Dashboard first, then analytical, then data, then hidden raw.
    order = ["Dashboard", "Daily", "Recovery", "Sleep", "Strain_Workouts",
             "Rolling", "Correlations", "Interventions", "Config",
             "Raw_Cycles", "Raw_Recovery", "Raw_Sleep", "Raw_Workouts"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    # Force a full recalc when opened in Excel so cached values are fresh.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    # Default Arial on the workbook's normal style is not directly settable;
    # every cell we wrote already uses Arial via explicit fonts where it shows.
    wb.save(out_path)

    if recalc:
        _recalc(out_path)

    _inject_sparklines(out_path, "Dashboard", spark_specs)

    errors = scan_for_errors(out_path)
    if errors:
        raise RuntimeError(
            "Workbook contains formula errors:\n  " + "\n  ".join(errors[:25])
        )
    log.info("Workbook written to %s (%d daily rows).", out_path, len(daily))
    return out_path
